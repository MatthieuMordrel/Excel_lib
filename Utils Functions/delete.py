from typing import List, Set, Optional, Literal
import openpyxl
from openpyxl.utils import range_boundaries
import os

# Default path and table configurations
DEFAULT_PATH = r"C:\Users\matth\OneDrive - Matthieu Mordrel\Work\Projects\Kovera\Project 2\Analysis of Files"
DEFAULT_FILE_NAME = "12. Grouped Data.xlsx"
DEFAULT_FULL_PATH = os.path.join(DEFAULT_PATH, DEFAULT_FILE_NAME)
DEFAULT_SHEET = "Sheet1"  # Assuming the default sheet name

# Table configurations
TABLE_CONFIGS = {
    "P_P": {
        "sheet_name": "P_P",
        "table_name": "P_P_Table",
        "column_name": "Product_Code_Parent"
    },
    "P_E": {
        "sheet_name": "P_E",
        "table_name": "P_E_Table",
        "column_name": "Product_Code"
    },
    "P_LADE": {
        "sheet_name": "P_LADE",
        "table_name": "P_LADE_Table",
        "column_name": "Product_Code"
    }
}

def delete_rows_by_values(
    values_to_delete: List[str],
    excel_path: Optional[str] = None,
    sheet_name: Optional[str] = None,
    table_name: Optional[str] = None,
    column_name: Optional[str] = None,
    save_changes: bool = True,
    table_type: Optional[Literal["P_P", "P_E", "P_LADE"]] = None
) -> int:
    """
    Deletes rows from an Excel table where values in a specified column match any value in the provided list.
    
    Args:
        values_to_delete (List[str]): List of values to search for and delete matching rows
        excel_path (str, optional): Path to the Excel file. Defaults to the standard analysis file.
        sheet_name (str, optional): Name of the worksheet containing the table. Defaults to "Sheet1".
        table_name (str, optional): Name of the table to search in. If table_type is provided, this is determined automatically.
        column_name (str, optional): Name of the column to search in. If table_type is provided, this is determined automatically.
        save_changes (bool, optional): Whether to save changes to the file. Defaults to True.
        table_type (str, optional): Shortcut to specify a predefined table configuration. Options: "P_P", "P_E", "P_LADE".
        
    Returns:
        int: Number of rows deleted
        
    Example:
        # Delete rows where Product Code is "P001" or "P002" in the P_E table
        delete_rows_by_values(
            ["P001", "P002"], 
            table_type="P_E"
        )
        
        # Delete rows with full custom parameters
        delete_rows_by_values(
            ["P001", "P002"], 
            "C:/path/to/file.xlsx", 
            "Products", 
            "ProductTable", 
            "ProductID"
        )
    """
    # Apply default configurations if table_type is provided
    if table_type and table_type in TABLE_CONFIGS:
        config = TABLE_CONFIGS[table_type]
        sheet_name = config["sheet_name"]
        table_name = config["table_name"]
        column_name = config["column_name"]
    
    # Apply other defaults
    if excel_path is None:
        excel_path = DEFAULT_FULL_PATH
    
    if sheet_name is None:
        sheet_name = DEFAULT_SHEET
    
    # Ensure all required parameters are provided
    if table_name is None:
        raise ValueError("Table name must be provided either directly or via table_type")
    
    if column_name is None:
        raise ValueError("Column name must be provided either directly or via table_type")
    
    # Track number of deleted rows
    deleted_count = 0
    
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(excel_path)
        
        # Get the worksheet
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        
        worksheet = workbook[sheet_name]
        
        # Find the table
        table = None
        for t in worksheet.tables.values():
            if t.name == table_name or t.displayName == table_name:
                table = t
                break
        
        if table is None:
            raise ValueError(f"Table '{table_name}' not found in sheet '{sheet_name}'")
        
        # Parse table range (e.g., "A1:D10")
        table_range = table.ref
        # Get the boundaries and ensure they are integers
        boundaries = range_boundaries(table_range)
        # The range_boundaries function returns a tuple of 4 values
        # We need to ensure they are all integers
        min_col, min_row, max_col, max_row = [
            int(val) if val is not None else 0 for val in boundaries
        ]
        
        # Get column headers and find the target column index
        header_row_idx = min_row
        column_index = None
        
        for col in range(min_col, max_col + 1):
            cell_value = worksheet.cell(row=header_row_idx, column=col).value
            if cell_value == column_name:
                column_index = col
                break
        
        if column_index is None:
            raise ValueError(f"Column '{column_name}' not found in table '{table_name}'")
        
        # Convert values_to_delete to a set for faster lookups (case-insensitive)
        values_set: Set[str] = {str(v).lower() for v in values_to_delete}
        
        # Find rows to delete (in reverse order to avoid index shifting issues)
        rows_to_delete: List[int] = []
        for row in range(max_row, header_row_idx, -1):
            cell_value = worksheet.cell(row=row, column=column_index).value
            
            # Skip None values
            if cell_value is None:
                continue
                
            # Convert to string and compare case-insensitively
            if str(cell_value).lower() in values_set:
                rows_to_delete.append(row)
                deleted_count += 1
        
        # Delete the identified rows
        for row in rows_to_delete:
            worksheet.delete_rows(row)
        
        # Save if requested
        if save_changes:
            workbook.save(excel_path)
            
        return deleted_count
        
    except Exception as e:
        raise e

def delete_from_pp_table(values_to_delete: List[str], excel_path: Optional[str] = None) -> int:
    """
    Deletes rows from the P_P_Table where Product_Code_Parent matches any value in the provided list.
    
    Args:
        values_to_delete (List[str]): List of Product_Code_Parent values to delete
        excel_path (str, optional): Path to the Excel file. Defaults to the standard analysis path.
        
    Returns:
        int: Number of rows deleted
    """
    return delete_rows_by_values(values_to_delete, excel_path=excel_path, table_type="P_P")

def delete_from_pe_table(values_to_delete: List[str], excel_path: Optional[str] = None) -> int:
    """
    Deletes rows from the P_E_Table where Product_Code matches any value in the provided list.
    
    Args:
        values_to_delete (List[str]): List of Product_Code values to delete
        excel_path (str, optional): Path to the Excel file. Defaults to the standard analysis path.
        
    Returns:
        int: Number of rows deleted
    """
    return delete_rows_by_values(values_to_delete, excel_path=excel_path, table_type="P_E")

def delete_from_plade_table(values_to_delete: List[str], excel_path: Optional[str] = None) -> int:
    """
    Deletes rows from the P_LADE_Table where Product_Code matches any value in the provided list.
    
    Args:
        values_to_delete (List[str]): List of Product_Code values to delete
        excel_path (str, optional): Path to the Excel file. Defaults to the standard analysis path.
        
    Returns:
        int: Number of rows deleted
    """
    return delete_rows_by_values(values_to_delete, excel_path=excel_path, table_type="P_LADE")

# =============================================================================
# CONFIGURATION SECTION - MODIFY THIS PART TO RUN THE SCRIPT
# =============================================================================

# Set this to True to run the script when this file is executed
RUN_SCRIPT = True

# Choose which operation to perform by setting exactly ONE of these to True
DELETE_FROM_PP_TABLE = False
DELETE_FROM_PE_TABLE = False
DELETE_FROM_PLADE_TABLE = True
DELETE_FROM_CUSTOM_TABLE = False

# Values to delete - MODIFY THIS LIST with your values
VALUES_TO_DELETE = [
    "CO2PBL120_10",
    "CO2PBL110_10", 
    "CO2PBL120_9",
    "CO2PBL110_9",
    "CO2PBL120_8",
    "CO2PBL120_7",
    "CO2PBL110_8",
    "CO2PBL120_6",
    "CO2PBL110_7",
    "CO2PBL110_6",
    "CO2PBL120_5",
    "CO2PBL110_5",
    "CO2PBL120_4",
    "CO2PBL110_4",
    "CO2PBL120_3",
    "CO2PBL110_3",
    "CO2PBL120_2",
    "CO2PBL110_2",
    "CO2PBL110_1",
    "CO2PBL120_1"
]

# -------------------------------------------------------------------------
# FILE PATH CONFIGURATION
# -------------------------------------------------------------------------
# By default, the script uses this file for all operations:
# C:\Users\matth\OneDrive - Matthieu Mordrel\Work\Projects\Kovera\Project 2\Analysis of Files\12. Grouped Data.xlsx
#
# To use a different file, set USE_DEFAULT_FILE to False and specify the full path below
USE_DEFAULT_FILE = True
CUSTOM_FILE_PATH = r"C:\Users\matth\OneDrive - Matthieu Mordrel\Work\Projects\Kovera\Project 2\Analysis of Files\Test_Delete.xlsx"

# Custom table parameters (only used if DELETE_FROM_CUSTOM_TABLE is True)
CUSTOM_SHEET_NAME = "Sheet1"
CUSTOM_TABLE_NAME = "Table1"
CUSTOM_COLUMN_NAME = "id"

# =============================================================================
# END OF CONFIGURATION SECTION
# =============================================================================

if __name__ == "__main__":
    if RUN_SCRIPT:
        try:
            # Determine which file path to use
            file_path = DEFAULT_FULL_PATH if USE_DEFAULT_FILE else CUSTOM_FILE_PATH
            
            if DELETE_FROM_PP_TABLE:
                # Delete from P_P_Table
                deleted = delete_from_pp_table(VALUES_TO_DELETE, excel_path=file_path)
                print(f"Successfully deleted {deleted} rows from P_P_Table where Product_Code_Parent matched any of: {', '.join(VALUES_TO_DELETE)}")
                
            elif DELETE_FROM_PE_TABLE:
                # Delete from P_E_Table
                deleted = delete_from_pe_table(VALUES_TO_DELETE, excel_path=file_path)
                print(f"Successfully deleted {deleted} rows from P_E_Table where Product_Code matched any of: {', '.join(VALUES_TO_DELETE)}")
                
            elif DELETE_FROM_PLADE_TABLE:
                # Delete from P_LADE_Table
                deleted = delete_from_plade_table(VALUES_TO_DELETE, excel_path=file_path)
                print(f"Successfully deleted {deleted} rows from P_LADE_Table where Product_Code matched any of: {', '.join(VALUES_TO_DELETE)}")
                
            elif DELETE_FROM_CUSTOM_TABLE:
                # Delete from custom table
                deleted = delete_rows_by_values(
                    values_to_delete=VALUES_TO_DELETE,
                    excel_path=file_path,
                    sheet_name=CUSTOM_SHEET_NAME,
                    table_name=CUSTOM_TABLE_NAME,
                    column_name=CUSTOM_COLUMN_NAME
                )
                print(f"Successfully deleted {deleted} rows from {CUSTOM_TABLE_NAME} where {CUSTOM_COLUMN_NAME} matched any of: {', '.join(VALUES_TO_DELETE)}")
            
            else:
                print("No operation selected. Please set one of the DELETE_* variables to True.")
                
        except Exception as e:
            print(f"Error: {e}")
    else:
        print("Script is disabled. Set RUN_SCRIPT to True to enable it.")
