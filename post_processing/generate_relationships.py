import json
from pathlib import Path
from typing import List, Dict, Any, Union
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

def load_processed_log(log_path: Path) -> List[Dict[str, Any]]:
    """Load and return the processed log data"""
    try:
        if not log_path.exists():
            raise FileNotFoundError(f"Processed log file not found: {log_path}")
        with open(log_path, 'r') as f:
            print(f"✅ Successfully loaded processed log from {log_path}")
            return json.load(f)
    except Exception as e:
        print(f"❌ Error loading processed log: {e}")
        raise

def process_product(product: Dict[str, Any], product_id: str, 
                    p_p_relations: List[Dict[str, Union[str, int]]],
                    p_e_relations: List[Dict[str, Union[str, int]]],
                    p_bm_relations: List[Dict[str, Union[str, int]]]) -> None:
    """
    Process a product and its references to extract relationships.
    Stops recursion when encountering a product, element, or base material.
    """
    try:
        for ref in product.get('references', []):
            if ref.get('isProduct'):
                # Check if relationship already exists
                existing_relation = next(
                    (rel for rel in p_p_relations 
                     if rel['Product_Id'] == product_id 
                     and rel['Product_Id_Child'] == ref.get('productID', '')), 
                    None
                )
                if existing_relation:
                    existing_relation['Quantity'] = int(existing_relation['Quantity']) + 1
                else:
                    p_p_relations.append({
                        'Product_Id': product_id,
                        'Product_Id_Child': ref.get('productID', ''),
                        'Quantity': 1
                    })
                print(f"➕ Added P_P relationship: {product_id} -> {ref.get('productID', '')}")
                continue
                
            if ref.get('isElement'):
                element_id = f"{ref.get('sheet', '')}_{round(ref.get('value', 0), 3)}"
                # Check if relationship already exists
                existing_relation = next(
                    (rel for rel in p_e_relations 
                     if rel['Product_Id'] == product_id 
                     and rel['Element_Id'] == element_id), 
                    None
                )
                if existing_relation:
                    existing_relation['Quantity'] = int(existing_relation['Quantity']) + 1
                else:
                    p_e_relations.append({
                        'Product_Id': product_id,
                        'Element_Id': element_id,
                        'Quantity': 1
                    })
                print(f"➕ Added P_E relationship: {product_id} -> {element_id}")
                continue
                
            if ref.get('isBaseMaterial'):
                bm_cell = f"{ref.get('sheet', '')}!{ref.get('cell', '')}"
                # Check if relationship already exists
                existing_relation = next(
                    (rel for rel in p_bm_relations 
                     if rel['Product_Id'] == product_id 
                     and rel['BM_Cell'] == bm_cell), 
                    None
                )
                if existing_relation:
                    existing_relation['Quantity'] = int(existing_relation['Quantity']) + 1
                else:
                    p_bm_relations.append({
                        'Product_Id': product_id,
                        'BM_Cell': bm_cell,
                        'Quantity': 1
                    })
                print(f"➕ Added P_BM relationship: {product_id} -> {bm_cell}")
                continue
                
            process_product(ref, product_id, p_p_relations, p_e_relations, p_bm_relations)
    except Exception as e:
        print(f"❌ Error processing product {product_id}: {e}")
        raise

def update_excel_sheet(sheet: Worksheet, data: List[Dict[str, str]], headers: List[str]) -> None:
    """
    Update an Excel sheet with new data while preserving existing content.
    Adds new rows below existing data.
    """
    try:
        start_row = sheet.max_row + 1 if sheet.max_row > 1 else 2
        print(f"📝 Updating sheet {sheet.title} starting at row {start_row}")
        
        for i, row in enumerate(data, start=start_row):
            for j, header in enumerate(headers, start=1):
                sheet.cell(row=i, column=j, value=row.get(header, ''))
        print(f"✅ Successfully updated {len(data)} rows in sheet {sheet.title}")
    except Exception as e:
        print(f"❌ Error updating sheet {sheet.title}: {e}")
        raise

def generate_relationships(log_path: Path, excel_path: Path) -> None:
    """
    Main function to generate relationships and update the Excel file.
    Creates a new sheet with the relationships tables.
    """
    try:
        print("\n🚀 Starting relationship generation process")
        
        # Load processed log data
        data = load_processed_log(log_path)
        print(f"📊 Found {len(data)} entries in processed log")
        
        # Initialize relationship containers
        p_p_relations: List[Dict[str, Union[str, int]]] = []
        p_e_relations: List[Dict[str, Union[str, int]]] = []
        p_bm_relations: List[Dict[str, Union[str, int]]] = []
        
        # Process each root object
        product_count = 0
        for item in data:
            if item.get('isProduct'):
                product_count += 1
                product_id = item.get('productID', '')
                print(f"\n🔍 Processing product {product_id}")
                process_product(item, product_id, p_p_relations, p_e_relations, p_bm_relations)
        
        print(f"\n📊 Found {product_count} products with relationships")
        print(f"  - Product-Product relationships: {len(p_p_relations)}")
        print(f"  - Product-Element relationships: {len(p_e_relations)}")
        print(f"  - Product-BaseMaterial relationships: {len(p_bm_relations)}")
        
        # Load existing Excel workbook
        try:
            workbook = openpyxl.load_workbook(excel_path)
            print(f"📂 Successfully loaded Excel workbook from {excel_path}")
        except Exception as e:
            print(f"❌ Error loading Excel workbook: {e}")
            raise
        
        # Create new sheet with timestamp
        from datetime import datetime
        new_sheet_name = f"Relationships_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        sheet = workbook.create_sheet(new_sheet_name)
        print(f"📄 Created new sheet: {new_sheet_name}")
        
        # Add tables with spacing between them
        current_row = 1
        
        # Add Product-Product table
        if p_p_relations:
            sheet.cell(row=current_row, column=1, value="Product-Product Relationships")
            current_row += 1
            sheet.cell(row=current_row, column=1, value="Product_Id")
            sheet.cell(row=current_row, column=2, value="Product_Id_Child")
            sheet.cell(row=current_row, column=3, value="Quantity")
            current_row += 1
            for relation in p_p_relations:
                sheet.cell(row=current_row, column=1, value=relation['Product_Id'])
                sheet.cell(row=current_row, column=2, value=relation['Product_Id_Child'])
                sheet.cell(row=current_row, column=3, value=relation['Quantity'])
                current_row += 1
            current_row += 2  # Add spacing between tables
        else:
            print("ℹ️ No Product-Product relationships found")
        
        # Add Product-Element table
        if p_e_relations:
            sheet.cell(row=current_row, column=1, value="Product-Element Relationships")
            current_row += 1
            sheet.cell(row=current_row, column=1, value="Product_Id")
            sheet.cell(row=current_row, column=2, value="Element_Id")
            sheet.cell(row=current_row, column=3, value="Quantity")
            current_row += 1
            for relation in p_e_relations:
                sheet.cell(row=current_row, column=1, value=relation['Product_Id'])
                sheet.cell(row=current_row, column=2, value=relation['Element_Id'])
                sheet.cell(row=current_row, column=3, value=relation['Quantity'])
                current_row += 1
            current_row += 2  # Add spacing between tables
        else:
            print("ℹ️ No Product-Element relationships found")
        
        # Add Product-BaseMaterial table
        if p_bm_relations:
            sheet.cell(row=current_row, column=1, value="Product-BaseMaterial Relationships")
            current_row += 1
            sheet.cell(row=current_row, column=1, value="Product_Id")
            sheet.cell(row=current_row, column=2, value="BM_Cell")
            sheet.cell(row=current_row, column=3, value="Quantity")
            current_row += 1
            for relation in p_bm_relations:
                sheet.cell(row=current_row, column=1, value=relation['Product_Id'])
                sheet.cell(row=current_row, column=2, value=relation['BM_Cell'])
                sheet.cell(row=current_row, column=3, value=relation['Quantity'])
                current_row += 1
        else:
            print("ℹ️ No Product-BaseMaterial relationships found")
        
        # Save the updated workbook
        try:
            workbook.save(excel_path)
            print(f"💾 Successfully saved updated workbook to {excel_path}")
        except Exception as e:
            print(f"❌ Error saving workbook: {e}")
            raise
        
        print("\n🎉 Relationship generation completed successfully!")
        
    except Exception as e:
        print(f"\n❌❌❌ Error in relationship generation: {e}")
        raise

if __name__ == "__main__":
    try:
        # Define paths
        log_path = Path(__file__).parent / "processed_log.json"
        excel_path = Path(__file__).parent / "Relationships.xlsx"
        
        print(f"📄 Log file path: {log_path}")
        print(f"📊 Excel file path: {excel_path}")
        
        # Generate and update relationships
        generate_relationships(log_path, excel_path)
    except Exception as e:
        print(f"\n❌❌❌ Fatal error: {e}") 