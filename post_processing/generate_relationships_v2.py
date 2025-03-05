import json
from pathlib import Path
from typing import List, Dict, Any, Union
import openpyxl
from datetime import datetime

def load_processed_log(log_path: Path) -> List[Dict[str, Any]]:
    """Load and return the processed log data"""
    try:
        if not log_path.exists():
            raise FileNotFoundError(f"Processed log file not found: {log_path}")
        with open(log_path, 'r') as f:
            print(f"✅ Successfully loaded processed log from {log_path}")
            return json.load(f)
    except Exception as e:
        print(f"❌ Error loading processed log: {e}")
        raise

def extract_relationships(data: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    """
    Extract relationships from the processed log data
    Returns a dictionary with three relationship types:
    - product_element: Product to Element relationships
    - product_product: Product to Product relationships
    - product_base_material: Product to Base Material relationships
    - product_binnenlade: Product to Binnenlade relationships
    - product_binnenpottenlade: Product to Binnenpottenlade relationships
    - product_hardcoded: Product to Hardcoded relationships
    """
    relationships: Dict[str, List[Dict[str, Union[str, int]]]] = {
        'product_element': [],
        'product_product': [],
        'product_base_material': [],
        'product_binnenlade': [],
        'product_binnenpottenlade': [],
        'product_hardcoded': []
    }

    for product in data:
        if product.get('type') != 'product':
            continue
            
        product_id = product.get('id', '')
        for ref in product.get('references', []):
            # Create relationship entry based on reference type
            relationship = {
                'product_id': product_id,
                'related_id': ref.get('id', ''),
                'quantity': ref.get('quantity', 1),
                'type': ref.get('type', '')
            }
            
            # Classify and store the relationship
            if ref.get('type') == 'element':
                relationships['product_element'].append(relationship)
            elif ref.get('type') == 'product':
                relationships['product_product'].append(relationship)
            elif ref.get('type') == 'baseMaterial':
                relationships['product_base_material'].append(relationship)
            elif ref.get('type') == 'binnenlade':
                relationships['product_binnenlade'].append(relationship)
            elif ref.get('type') == 'binnenpottenlade':
                relationships['product_binnenpottenlade'].append(relationship)
            elif ref.get('type') == 'hardcoded':
                relationships['product_hardcoded'].append(relationship)
                
    return relationships

def create_relationship_sheet(workbook: openpyxl.Workbook, 
                             relationships: Dict[str, List[Dict[str, Any]]]) -> None:
    """
    Create a new sheet in the workbook with the relationship data
    """
    sheet_name = f"Relationships_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    sheet = workbook.create_sheet(sheet_name)
    current_row = 1
    
    # Helper function to add a relationship table
    def add_relationship_table(title: str, data: List[Dict[str, Any]], columns: List[str]):
        nonlocal current_row
        if not data:
            return
            
        # Add table header
        sheet.cell(row=current_row, column=1, value=title)
        current_row += 1
        
        # Add column headers
        for col_idx, col_name in enumerate(columns, start=1):
            sheet.cell(row=current_row, column=col_idx, value=col_name)
        current_row += 1
        
        # Add data rows
        for relationship in data:
            for col_idx, col_name in enumerate(columns, start=1):
                sheet.cell(row=current_row, column=col_idx, value=relationship.get(col_name, ''))
            current_row += 1
        
        current_row += 2  # Add spacing between tables
    
    # Add Product-Element relationships
    add_relationship_table(
        "Product-Element Relationships",
        relationships['product_element'],
        ['product_id', 'related_id', 'quantity']
    )
    
    # Add Product-Product relationships
    add_relationship_table(
        "Product-Product Relationships",
        relationships['product_product'],
        ['product_id', 'related_id', 'quantity']
    )
    
    # Add Product-BaseMaterial relationships
    add_relationship_table(
        "Product-BaseMaterial Relationships",
        relationships['product_base_material'],
        ['product_id', 'related_id', 'quantity']
    )
    
    # Add Product-Binnenlade relationships
    add_relationship_table(
        "Product-Binnenlade Relationships",
        relationships['product_binnenlade'],
        ['product_id', 'related_id', 'quantity']
    )
    
    # Add Product-Binnenpottenlade relationships
    add_relationship_table(
        "Product-Binnenpottenlade Relationships",
        relationships['product_binnenpottenlade'],
        ['product_id', 'related_id', 'quantity']
    )

    # Add Product-Hardcoded relationships
    add_relationship_table(
        "Product-Hardcoded Relationships",
        relationships['product_hardcoded'],
        ['product_id', 'related_id', 'quantity']
    )

def generate_relationships_v2(log_path: Path, excel_path: Path) -> None:
    """
    Main function to generate relationships from processed_log.json format
    and save them to an Excel file
    """
    try:
        print("\n🚀 Starting relationship generation process")
        
        # Load processed log data
        data = load_processed_log(log_path)
        print(f"📊 Found {len(data)} entries in processed log")
        
        # Extract relationships
        relationships = extract_relationships(data)
        print(f"📊 Extracted relationships:")
        print(f"  - Product-Element: {len(relationships['product_element'])}")
        print(f"  - Product-Product: {len(relationships['product_product'])}")
        print(f"  - Product-BaseMaterial: {len(relationships['product_base_material'])}")
        print(f"  - Product-Binnenlade: {len(relationships['product_binnenlade'])}")
        print(f"  - Product-Binnenpottenlade: {len(relationships['product_binnenpottenlade'])}")
        print(f"  - Product-Hardcoded: {len(relationships['product_hardcoded'])}")
        
        # Create or load Excel workbook
        if excel_path.exists():
            workbook = openpyxl.load_workbook(excel_path)
            print(f"📂 Loaded existing Excel workbook: {excel_path}")
        else:
            workbook = openpyxl.Workbook()
            print(f"📂 Created new Excel workbook: {excel_path}")
        
        # Add relationship data to workbook
        create_relationship_sheet(workbook, relationships)
        
        # Save the workbook
        workbook.save(excel_path)
        print(f"💾 Saved relationships to {excel_path}")
        
        print("\n🎉 Relationship generation completed successfully!")
        
    except Exception as e:
        print(f"\n❌❌❌ Error in relationship generation: {e}")
        raise

if __name__ == "__main__":
    try:
        # Define paths
        path = 'Logs/Current Logs'
        log_path = Path(path) / "processed_log.json"
        excel_path = Path(path) / "Relationships_v2.xlsx"
        
        # Generate and update relationships
        generate_relationships_v2(log_path, excel_path)
    except Exception as e:
        print(f"\n❌❌❌ Fatal error: {e}") 